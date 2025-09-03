import os
import sys
import glob
import pandas as pd
import matplotlib.pyplot as plt
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# ---------------- Machine Mapping ----------------
MACHINE_MAP = {
    "400706": "M1",
    "400697": "M2",
    "400687": "M3",
    "400700": "M4",
    "400709": "M5",
    "4001010": "M6",
    "400845": "M7",
    "4001015": "M8",
}

# ---------------- Helper Functions ----------------
def parse_log_file(filepath):
    record = {"file": os.path.basename(filepath)}
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()
    for line in lines:
        if line.startswith("StartTime:"):
            record["start_time"] = pd.to_datetime(line.split("\t")[1].strip())
        elif line.startswith("EndTime:"):
            record["end_time"] = pd.to_datetime(line.split("\t")[1].strip())
        elif line.startswith("PrintingPeriod:"):
            period_str = line.split("\t")[1].strip()
            d, hms = period_str.split(".")
            h, m, s = map(int, hms.split(":"))
            td = timedelta(days=int(d), hours=h, minutes=m, seconds=s)
            record["runtime_sec"] = td.total_seconds()
    record["machine"] = infer_machine_name(filepath)
    return record

def infer_machine_name(filepath):
    fname = os.path.basename(filepath)
    for key, machine in MACHINE_MAP.items():
        if key in fname:
            return machine
    print(f"[WARNING] Unknown machine for file: {fname}")
    return "Unknown"

def generate_summary(filepaths, scheduled_time=1080):
    records = [parse_log_file(fp) for fp in filepaths]
    df = pd.DataFrame(records)
    df.drop_duplicates(inplace=True)
    df["date"] = df["start_time"].dt.date
    df["runtime_min"] = df["runtime_sec"] / 60

    summary = df.groupby(["date", "machine"]).agg(
        total_runtime_min=("runtime_min", "sum"),
        job_count=("runtime_min", "count"),
        avg_runtime_min=("runtime_min", "mean")
    ).reset_index()

    summary["utilisation_pct"] = (summary["total_runtime_min"] / scheduled_time) * 100
    return df, summary

def create_excel_report(summary, raw_df, output_xlsx="Daily_Report.xlsx"):
    wb = Workbook()

    # Raw Data Sheet
    ws1 = wb.active
    ws1.title = "Raw Data"
    ws1.append(list(raw_df.columns))
    for row in raw_df.itertuples(index=False):
        ws1.append(list(row))

    # Summary Sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(list(summary.columns))
    for row in summary.itertuples(index=False):
        ws2.append(list(row))

    # Dashboard Sheet
    ws3 = wb.create_sheet("Dashboard")
    ws3.append(["Machine", "Utilisation %"])
    latest_date = summary["date"].max()
    daily = summary[summary["date"] == latest_date]

    for _, row in daily.iterrows():
        ws3.append([row["machine"], row["utilisation_pct"]])

    chart = BarChart()
    chart.title = f"Machine Utilisation % - {latest_date}"
    chart.y_axis.title = "Utilisation %"
    chart.x_axis.title = "Machine"
    data = Reference(ws3, min_col=2, min_row=1, max_row=len(daily)+1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(daily)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws3.add_chart(chart, "E2")

    wb.save(output_xlsx)
    return output_xlsx

# ---------------- Main Script ----------------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python daily_utilisation.py <log_folder>")
        sys.exit(1)

    folder = sys.argv[1]
    files = glob.glob(os.path.join(folder, "*.txt"))

    if not files:
        print("No log files found in folder.")
        sys.exit(1)

    raw_df, summary = generate_summary(files, scheduled_time=1080)
    latest_date = summary["date"].max()
    out_xlsx = f"Daily_Report_{latest_date}.xlsx"
    create_excel_report(summary, raw_df, output_xlsx=out_xlsx)

    print(f"Report generated: {out_xlsx}")
